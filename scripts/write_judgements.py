import argparse
import datetime as dt
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


DEFAULT_FIELD_MAP = {
    "客户类型判断": "customer_type",
    "适合产品方向": "product_direction",
    "匹配理由": "reason",
    "开发切入点": "cut_in",
    "官网验证状态": "website_status",
    "官网验证摘要": "website_summary",
    "不适合原因": "unsuitable_reason",
    "社媒验证状态": "social_status",
    "社媒活跃度": "social_activity",
    "社媒近6个月业务信号": "social_signal",
    "社媒匹配判断": "social_judgement",
    "社媒辅助评分调整": "social_adjust",
    "社媒验证摘要": "social_signal",
    "综合匹配评分": "score",
    "综合开发建议": "development_advice",
}


def normalize(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip()).casefold()


def expected_level(score: int) -> str:
    if score >= 80:
        return "高"
    if score >= 60:
        return "中"
    if score >= 40:
        return "低"
    return "不适合"


def headers(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-xlsx", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--judgements-json", required=True)
    parser.add_argument("--level-column", required=True, help="Excel column containing the final fit level")
    parser.add_argument("--marker", default="【Agent证据判断】")
    parser.add_argument("--start", type=int)
    parser.add_argument("--count", type=int)
    args = parser.parse_args()

    target = Path(args.target_xlsx)
    items = json.loads(Path(args.judgements_json).read_text(encoding="utf-8"))
    if args.start:
        items = items[args.start - 1 :]
    if args.count:
        items = items[: args.count]
    if not items:
        raise RuntimeError("No judgement items selected")

    for item in items:
        score = item.get("score")
        if not isinstance(score, int) or not 0 <= score <= 100:
            raise RuntimeError(f"{item.get('company')}: invalid score {score!r}")
        if item.get("level") != expected_level(score):
            raise RuntimeError(f"{item.get('company')}: level does not match score")
        if item.get("social_adjust") not in {0, 3, 5}:
            raise RuntimeError(f"{item.get('company')}: invalid social_adjust")
        if item.get("level") in {"低", "不适合"} and not str(item.get("unsuitable_reason", "")).strip():
            raise RuntimeError(f"{item.get('company')}: missing unsuitable_reason")

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = target.with_name(f"{target.stem}.before_write_agent_{timestamp}{target.suffix}")
    shutil.copy2(target, backup)

    wb = load_workbook(target)
    ws = wb[args.sheet]
    hm = headers(ws)
    if "Company Name" not in hm:
        raise RuntimeError("Missing Company Name column")
    if args.level_column not in hm:
        raise RuntimeError(f"Missing level column: {args.level_column}")

    field_map = dict(DEFAULT_FIELD_MAP)
    field_map[args.level_column] = "level"
    if "匹配评分" in hm:
        field_map["匹配评分"] = "score"
    if "数字标牌匹配层级" in hm:
        field_map["数字标牌匹配层级"] = "level"
    if "Scanner匹配层级" in hm:
        field_map["Scanner匹配层级"] = "level"
    if "POS匹配层级" in hm:
        field_map["POS匹配层级"] = "level"

    missing = [col for col in field_map if col not in hm]
    field_map = {col: key for col, key in field_map.items() if col in hm}

    by_company = {normalize(item["company"]): item for item in items}
    updated = []
    for row_idx in range(2, ws.max_row + 1):
        company = str(ws.cell(row_idx, hm["Company Name"]).value or "")
        item = by_company.get(normalize(company))
        if not item:
            continue
        if "公司简介信息" in hm:
            intro = str(ws.cell(row_idx, hm["公司简介信息"]).value or "")
            cut = intro.find("【Agent证据判断")
            if cut >= 0:
                intro = intro[:cut].rstrip()
            add = f"{args.marker}{item.get('website_summary','')}\n【Agent业务判断】{item.get('reason','')}"
            ws.cell(row_idx, hm["公司简介信息"]).value = f"{intro}\n\n{add}" if intro else add
        for col, key in field_map.items():
            cell = ws.cell(row_idx, hm[col])
            cell.value = item.get(key, "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        updated.append(company)

    if len(updated) != len(items):
        raise RuntimeError(f"Updated {len(updated)} rows, expected {len(items)}")
    ws.freeze_panes = None

    try:
        wb.save(target)
        output = target
    except PermissionError:
        output = target.with_name(f"{target.stem}.agent_write_{timestamp}{target.suffix}")
        wb.save(output)

    print(json.dumps({"updated": len(updated), "backup": str(backup), "output": str(output), "missing_optional_columns": missing}, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
