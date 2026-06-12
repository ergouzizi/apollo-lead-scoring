import argparse
import hashlib
import json
import re
import subprocess
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook


DEFAULT_ANYSEARCH_CLI = Path.home() / ".codex" / "skills" / "anysearch" / "scripts" / "anysearch_cli.py"


def clean_text(text: str) -> str:
    text = text or ""
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clip(text: str, limit: int) -> str:
    text = clean_text(text)
    return text if len(text) <= limit else text[:limit].rstrip() + "..."


def slug(text: str) -> str:
    digest = hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:10]
    safe = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")[:50]
    return f"{safe}_{digest}" if safe else digest


def normalize_url(url: str) -> str:
    url = (url or "").strip()
    if url and not re.match(r"^https?://", url, flags=re.I):
        url = "http://" + url
    return url


def domain_of(url: str) -> str:
    try:
        return urlparse(normalize_url(url)).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def same_domain(url: str, domain: str) -> bool:
    host = domain_of(url)
    return bool(domain and (host == domain or host.endswith("." + domain)))


def is_quota_error(text: str) -> bool:
    low = (text or "").lower()
    return any(
        marker in low
        for marker in [
            "quota exceeded",
            "quota exhausted",
            "rate limit",
            "too many requests",
            "auto_registered",
            "payment required",
            "insufficient quota",
            "insufficient credits",
            "limit exceeded",
        ]
    )


def run_anysearch(python_exe: str, cli: Path, args: list[str], timeout: int) -> dict:
    started = time.time()
    proc = subprocess.run(
        [python_exe, str(cli), *args],
        text=True,
        encoding="utf-8",
        errors="replace",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=timeout,
    )
    stdout = clean_text(proc.stdout)
    stderr = clean_text(proc.stderr)
    quota = is_quota_error(stdout + "\n" + stderr)
    return {
        "ok": proc.returncode == 0 and not quota,
        "returncode": proc.returncode,
        "seconds": round(time.time() - started, 2),
        "stdout": stdout,
        "stderr": stderr,
        "quota_or_api_error": quota,
    }


def candidate_urls(markdown: str, home_url: str) -> list[str]:
    domain = domain_of(home_url)
    urls = []
    for raw in re.findall(r"https?://[^\s\]\)\"']+", markdown or ""):
        url = raw.rstrip(".,;:)")
        if same_domain(url, domain):
            urls.append(url)
    for href in re.findall(r"\]\(([^)]+)\)", markdown or ""):
        if href.startswith("#") or href.lower().startswith(("mailto:", "tel:", "javascript:", "data:")):
            continue
        url = urljoin(home_url, href)
        if same_domain(url, domain):
            urls.append(url)
    out, seen = [], set()
    for url in urls:
        key = url.rstrip("/")
        if key not in seen:
            seen.add(key)
            out.append(url)
    return out


def pick_supplements(markdown: str, home_url: str, limit: int) -> list[str]:
    buckets = [
        ["about", "company", "profile", "gioi-thieu", "ve-chung-toi"],
        ["product", "products", "service", "services", "solution", "solutions", "san-pham", "dich-vu", "giai-phap"],
        ["cms", "digital-signage", "signage", "scanner", "barcode", "auto-id", "pos"],
    ]
    urls = candidate_urls(markdown, home_url)
    selected, seen = [], {home_url.rstrip("/")}
    for patterns in buckets:
        for url in urls:
            if url.rstrip("/") in seen:
                continue
            if any(p in url.lower() for p in patterns):
                selected.append(url)
                seen.add(url.rstrip("/"))
                if len(selected) >= limit:
                    return selected
    return selected


def headers(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-xlsx", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--prefix", default="run")
    parser.add_argument("--product-query", default="digital signage CMS screen management scanner POS")
    parser.add_argument("--start", type=int)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--supplement-limit", type=int, default=3)
    parser.add_argument("--python", default="python")
    parser.add_argument("--anysearch-cli", default=str(DEFAULT_ANYSEARCH_CLI))
    args = parser.parse_args()

    target = Path(args.target_xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = out_dir / "raw"
    raw_dir.mkdir(exist_ok=True)

    wb = load_workbook(target, read_only=True, data_only=True)
    ws = wb[args.sheet]
    hm = headers(ws)
    required = ["Company Name", "Website URL", "All Social Links", "Batch No.", "Industry", "Keywords", "Source File"]
    missing = [h for h in required if h not in hm]
    if missing:
        raise RuntimeError(f"Missing headers: {missing}")

    rows = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        company = row[hm["Company Name"] - 1]
        if not company:
            continue
        rows.append(
            {
                "excel_row": excel_row,
                "company": str(company),
                "website": row[hm["Website URL"] - 1] or "",
                "social_links": row[hm["All Social Links"] - 1] or "",
                "batch_no": row[hm["Batch No."] - 1] or "",
                "industry": row[hm["Industry"] - 1] or "",
                "keywords": row[hm["Keywords"] - 1] or "",
                "source_file": row[hm["Source File"] - 1] or "",
                "intro": row[hm.get("公司简介信息", 0) - 1] if hm.get("公司简介信息") else "",
            }
        )
    if args.start:
        rows = rows[args.start - 1 :]
    if args.limit:
        rows = rows[: args.limit]

    cli = Path(args.anysearch_cli)
    packs = []
    for idx, row in enumerate(rows, 1):
        print(json.dumps({"progress": f"{idx}/{len(rows)}", "company": row["company"]}, ensure_ascii=True), flush=True)
        cdir = raw_dir / slug(row["company"])
        cdir.mkdir(exist_ok=True)
        cached = cdir / "evidence_pack.json"
        if cached.exists():
            packs.append(json.loads(cached.read_text(encoding="utf-8")))
            continue

        website = normalize_url(str(row["website"]))
        homepage = {"ok": False, "stdout": "", "stderr": "no website", "quota_or_api_error": False}
        supplements = []
        if website:
            homepage = run_anysearch(args.python, cli, ["extract", website], 85)
            (cdir / "homepage.md").write_text(homepage["stdout"] or homepage["stderr"], encoding="utf-8")
            if homepage["quota_or_api_error"]:
                raise RuntimeError(f"AnySearch quota/api error on {row['company']}")
            if homepage["ok"]:
                for url in pick_supplements(homepage["stdout"], website, args.supplement_limit):
                    result = run_anysearch(args.python, cli, ["extract", url], 85)
                    (cdir / f"supplement_{len(supplements)+1}.md").write_text(result["stdout"] or result["stderr"], encoding="utf-8")
                    if result["quota_or_api_error"]:
                        raise RuntimeError(f"AnySearch quota/api error on {row['company']} supplement {url}")
                    supplements.append({"url": url, "ok": result["ok"], "content": result["stdout"]})

        query = f'"{row["company"]}" {args.product_query}'
        search = run_anysearch(args.python, cli, ["search", query, "--max_results", "5"], 75)
        (cdir / "search.md").write_text(search["stdout"] or search["stderr"], encoding="utf-8")
        if search["quota_or_api_error"]:
            raise RuntimeError(f"AnySearch quota/api error on search for {row['company']}")

        pack = {
            **row,
            "homepage_evidence_clip": clip(homepage["stdout"], 900),
            "supplement_evidence_clip": clip("\n\n".join(f"{s['url']}\n{s['content']}" for s in supplements), 1200),
            "search_evidence_clip": clip(search["stdout"], 900),
            "anysearch_status": {
                "homepage_ok": homepage["ok"],
                "search_ok": search["ok"],
                "supplement_count": sum(1 for s in supplements if s["ok"]),
            },
        }
        cached.write_text(json.dumps(pack, ensure_ascii=False, indent=2), encoding="utf-8")
        packs.append(pack)

    (out_dir / f"{args.prefix}_evidence_packs.json").write_text(json.dumps(packs, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {
        "count": len(packs),
        "homepage_ok": sum(1 for p in packs if p["anysearch_status"]["homepage_ok"]),
        "search_ok": sum(1 for p in packs if p["anysearch_status"]["search_ok"]),
        "supplement_pages": sum(p["anysearch_status"]["supplement_count"] for p in packs),
    }
    (out_dir / f"{args.prefix}_evidence_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
