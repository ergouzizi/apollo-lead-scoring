import argparse
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def expected_level(score: int) -> str:
    if score >= 80:
        return "高"
    if score >= 60:
        return "中"
    if score >= 40:
        return "低"
    return "不适合"


def headers(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-xlsx", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--level-column", required=True)
    parser.add_argument("--score-column", default="综合匹配评分")
    parser.add_argument("--unsuitable-column", default="不适合原因")
    parser.add_argument("--social-adjust-column", default="社媒辅助评分调整")
    parser.add_argument("--expect-freeze-none", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(Path(args.target_xlsx), data_only=True)
    ws = wb[args.sheet]
    hm = headers(ws)
    for col in ["Company Name", args.level_column, args.score_column, args.unsuitable_column, args.social_adjust_column]:
        if col not in hm:
            raise RuntimeError(f"Missing required column: {col}")

    errors = []
    distribution = Counter()
    rows = 0
    for row_idx in range(2, ws.max_row + 1):
        company = ws.cell(row_idx, hm["Company Name"]).value
        if not company:
            continue
        rows += 1
        score = ws.cell(row_idx, hm[args.score_column]).value
        level = ws.cell(row_idx, hm[args.level_column]).value
        distribution[level] += 1
        if not isinstance(score, int) or not 0 <= score <= 100:
            errors.append(f"{company}: invalid score {score!r}")
        elif level != expected_level(score):
            errors.append(f"{company}: level {level!r} does not match score {score}")
        social_adjust = ws.cell(row_idx, hm[args.social_adjust_column]).value
        if social_adjust not in {0, 3, 5}:
            errors.append(f"{company}: invalid social adjust {social_adjust!r}")
        unsuitable = str(ws.cell(row_idx, hm[args.unsuitable_column]).value or "").strip()
        if level in {"低", "不适合"} and not unsuitable:
            errors.append(f"{company}: low/unsuitable without unsuitable reason")
    if args.expect_freeze_none and ws.freeze_panes is not None:
        errors.append(f"freeze_panes is {ws.freeze_panes!r}, expected None")

    print(json.dumps({"rows": rows, "distribution": dict(distribution), "freeze_panes": ws.freeze_panes, "errors": errors[:100], "error_count": len(errors)}, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
